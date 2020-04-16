

import wx
import orr
import xlsxwriter
import openpyxl
from orr import GoogleDriveApi
import os
import Const
import pandas as pd
import xlrd


from shutil import copy2


import pandas as pd

from openpyxl import load_workbook
from openpyxl import load_workbook
class Sheet:
    def __init__(self,path:str):

        self.path=path
        self.workbook = openpyxl.load_workbook(self.path)
        print("yes!!")


    def addNewWorkSheet(self, data:pd.DataFrame, name_of_work:str):

        wb2 = load_workbook(self.path)
        wb2.create_sheet(str(name_of_work))



        wb2.save(self.path)

        writer = pd.ExcelWriter(self.path, engine='openpyxl')

        # try to open an existing workbook

        writer.book = load_workbook(self.path)

        # copy existing sheets

        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        # read existing file

        reader = pd.read_excel(str(self.path), sheet_name=str(name_of_work))

        # write out the new sheet

        data.to_excel(writer, index=False, sheet_name=str(name_of_work), header=True)

        writer.close()






    def GetFolderID(self,dic,courseId:str):
        pass


    def UpdateWorkSheetHWSOL(self,data:{},courseId:str,nameOfWork:str):
       # folderId,fileId,year,semster,lectureName,HWORSolution,numOfHW,partOfHW,pathToFile,remarks

       data_HW = pd.DataFrame({Const.FOLDERID:[data[Const.PATH_TO_FILE]],
                  Const.FILEID:[data[Const.FILEID]],
                  Const.YEAR: [data[Const.YEAR]],
                  Const.SEMSTER:[data[Const.SEMSTER]],
                  Const.LECTURE_NAME: [data[Const.LECTURE_NAME]],
                  Const.SOL_OR_HW: [data[Const.SOL_OR_HW]],
                  Const.NUM_OF_HW: [data[Const.NUM_OF_HW]],
                  Const.PART_OF_HW: [data[Const.PART_OF_HW]],
                  Const.PATH_TO_FILE_DRIVE: [data[Const.PATH_TO_FILE]],
                  Const.REMARKS: [data[Const.REMARKS]]
                  })

       writer = pd.ExcelWriter(self.path, engine='openpyxl')

        # try to open an existing workbook

       writer.book = load_workbook(self.path)

        # copy existing sheets

       writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        # read existing file

       reader = pd.read_excel(str(self.path), sheet_name=nameOfWork)

        # write out the new sheet

       data_HW.to_excel(writer, index=False, sheet_name=nameOfWork, header=False, startrow=len(reader) + 1)

       writer.close()

    def UpdateWorkSheetTest(self,data:{},courseId:str,nameOfWork:str):
        data[Const.FILEID] = 1
        data_TEST = pd.DataFrame({Const.FOLDERID: [data[Const.PATH_TO_FILE]],
                                    Const.FILEID: [data[Const.FILEID]],
                                    Const.YEAR: [data[Const.YEAR]],
                                    Const.SEMSTER: [data[Const.SEMSTER]],
                                    Const.LECTURE_NAME: [data[Const.LECTURE_NAME]],
                                    Const.MOED: [data[Const.MOED]],
                                    Const.TEST_OR_MIDTEST: [data[Const.TEST_OR_MIDTEST]],
                                    Const.QUE_OR_SOL: [data[Const.QUE_OR_SOL]],
                                    Const.PATH_TO_FILE_DRIVE: [data[Const.PATH_TO_FILE]],
                                    Const.REMARKS: [data[Const.REMARKS]]
                                    })

        writer = pd.ExcelWriter(self.path, engine='openpyxl')

        # try to open an existing workbook

        writer.book = load_workbook(self.path)

        # copy existing sheets

        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        # read existing file

        reader = pd.read_excel(str(self.path), sheet_name=nameOfWork)

        # write out the new sheet

        data_TEST.to_excel(writer, index=False, sheet_name=nameOfWork, header=False, startrow=len(reader) + 1)

        writer.close()


        pass

    def UpdateWorkSheetHelpStaff(self,data:{},courseId:str,nameOfWork:str):
        data[Const.FILEID] = 1
        data_helpStaff = pd.DataFrame({Const.FOLDERID: [data[Const.PATH_TO_FILE]],
                                    Const.FILEID: [data[Const.FILEID]],
                                    Const.YEAR: [data[Const.YEAR]],
                                    Const.SEMSTER: [data[Const.SEMSTER]],
                                    Const.WHAT_THIS_FILE: [data[Const.WHAT_THIS_FILE]],
                                    Const.REMARKS: [data[Const.REMARKS]]
                                    })

        writer = pd.ExcelWriter(self.path, engine='openpyxl')

        # try to open an existing workbook

        writer.book = load_workbook(self.path)

        # copy existing sheets

        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        # read existing file

        reader = pd.read_excel(str(self.path), sheet_name=nameOfWork)

        # write out the new sheet

        data_helpStaff.to_excel(writer, index=False, sheet_name=nameOfWork, header=False, startrow=len(reader) + 1)

        writer.close()

        pass

    def UpdateWorkSheetTOTLEC(self,data:{},courseId:str,nameOfWork:str):
       # folderId,fileId,year,semster,lectureName,HWORSolution,numOfHW,partOfHW,pathToFile,remarks
       data[Const.FILEID]=1

       data_TOTLEC = pd.DataFrame({Const.FOLDERID:[data[Const.PATH_TO_FILE]],
                  Const.FILEID:[data[Const.FILEID]],
                  Const.YEAR: [data[Const.YEAR]],
                  Const.SEMSTER:[data[Const.SEMSTER]],
                  Const.LECTURE_NAME: [data[Const.LECTURE_NAME]],
                  Const.SOL_OR_HW: [data[Const.LECTURE_OR_TOTURIAL]],
                  Const.NUM_OF_HW: [data[Const.NUMOFLECTURE]],
                  Const.PART_OF_HW: [data[Const.PARTOFLECTURE]],
                  Const.PATH_TO_FILE_DRIVE: [data[Const.PATH_TO_FILE]],
                  Const.REMARKS: [data[Const.REMARKS]]
                  })

       writer = pd.ExcelWriter(self.path, engine='openpyxl')

        # try to open an existing workbook

       writer.book = load_workbook(self.path)

        # copy existing sheets

       writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        # read existing file

       reader = pd.read_excel(str(self.path), sheet_name=nameOfWork)

        # write out the new sheet

       data_TOTLEC.to_excel(writer, index=False, sheet_name=nameOfWork, header=False, startrow=len(reader) + 1)

       writer.close()





    def InsertCourse(self,FolderPath:str,courseId:int,courseName:str):




        # new dataframe with same columns

        df = pd.DataFrame({'FolderPath': [FolderPath],

                           'CourseID': [courseId],
                           'CourseName': [courseName]

                           }
                          )

        writer = pd.ExcelWriter(self.path, engine='openpyxl')

        # try to open an existing workbook

        writer.book = load_workbook(self.path)

        # copy existing sheets

        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        # read existing file


        reader = pd.read_excel(str(self.path),sheet_name="Sheet1")

        # write out the new sheet

        df.to_excel(writer, index=False,sheet_name="Sheet1", header=False, startrow=len(reader) + 1)

        writer.close()




       # self.addNewWorkSheet(0,"yes")

       # writer = pd.ExcelWriter(self.workbook["Sheet1"],engine='openpyxl')
       # writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
       # print(writer.sheets)

        pass

    def FindDiractoryID(self,CourseId,folderName):

        pass



    def FindRootDiractoryID(self,CourseId:str):


        pass

    def FindDiractoryPathTest(self,courseId:str,dic:{}):
        work=str(courseId)+Const.FOLDERID
        reader = pd.read_excel(str(self.path), sheet_name=work)
        if dic[Const.TEST_OR_MIDTEST]==Const.TEST:
           if dic[Const.QUE_OR_SOL]==Const.QUE:
                one_drive_id_col = reader[reader['FolderName'] ==str("TestWithOutSolFolder")]
           else:
                one_drive_id_col = reader[reader['FolderName'] == str("TestWithSolFolder")]
        else:
            if dic[Const.QUE_OR_SOL] == Const.QUE:
                one_drive_id_col = reader[reader['FolderName'] == str("MidTestWithOutFolder")]
            else:
                one_drive_id_col = reader[reader['FolderName'] == str("MidTestWithFolder")]

        folder_path = one_drive_id_col.iloc[0, 1]
        print(folder_path)
        return folder_path



        pass

    def FindDiractoryPathHW(self,courseId:str,dic:{}):
        work = str(courseId) + Const.FOLDERID
        reader = pd.read_excel(str(self.path), sheet_name=work)
        one_drive_id_col = reader[reader['FolderName'] == Const.HWSOLUTION+str("Folder")]
        folder_path = one_drive_id_col.iloc[0, 1]
        print(folder_path)
        return folder_path

    def FindDiractoryPathToutrialLecture(self,courseId:str,dic:{}):
        work = str(courseId) + Const.FOLDERID
        reader = pd.read_excel(str(self.path), sheet_name=work)
        one_drive_id_col = reader[reader['FolderName'] == str(dic[Const.LECTURE_OR_TOTURIAL])+str("Folder")]
        folder_path = one_drive_id_col.iloc[0, 1]
        print(folder_path)
        return folder_path


    def FindDiractoryPathHelpStaff(self,courseId:str,dic:{}):
        work = str(courseId) + Const.FOLDERID
        reader = pd.read_excel(str(self.path), sheet_name=work)
        one_drive_id_col = reader[reader['FolderName'] == str("HelpStaffFolder")]
        folder_path = one_drive_id_col.iloc[0, 1]
        print(folder_path)
        return folder_path






class OneDriveApi:
    def __init__(self,path):
        self.working_dirctory =path
        self.mySheet =Sheet(str(path)+str({}).format("/Courses.xlsx"))



    def AddLectureToturial(self,data:{}):

        folderPath = self.mySheet.FindDiractoryPathToutrialLecture(data[Const.COURSE_ID], data)
        copy2(data[Const.PATH_TO_FILE], folderPath)

        name = "{} Number {} Part {} Semster {} Year {} LectureName {} Autor by {}".format(
            data[Const.LECTURE_OR_TOTURIAL], data[Const.NUMOFLECTURE],
            data[Const.PARTOFLECTURE], data[Const.SEMSTER]
            , data[Const.YEAR], data[Const.LECTURE_NAME], data[Const.REMARKS])

        my_name_list = data[Const.PATH_TO_FILE].split("\\")
        my_name = my_name_list[len(my_name_list) - 1]
        my_name_format = my_name.split('.')[1]

        newName = str(folderPath) + str('/') + str(name) + str('.') + my_name_format
        os.rename(str(folderPath) + str('/') + str(my_name), newName)
        data[Const.PATH_TO_FILE] = newName

        self.mySheet.UpdateWorkSheetTOTLEC(data=data, courseId=data[Const.COURSE_ID],
                                     nameOfWork=data[Const.COURSE_ID]+Const.Toturial_LECTURE_WORK)


    def AddTest(self,data:{}):
        folderPath = self.mySheet.FindDiractoryPathTest(data[Const.COURSE_ID], data)
        copy2(data[Const.PATH_TO_FILE], folderPath)
        name = "{} for {}  Year {}  Semster {} Moed {} LectureName {} Autor by {}".format(
            data[Const.QUE_OR_SOL],data[Const.TEST_OR_MIDTEST],data[Const.YEAR],
            data[Const.SEMSTER],data[Const.MOED],data[Const.LECTURE_NAME],
            data[Const.REMARKS])

        my_name_list = data[Const.PATH_TO_FILE].split("\\")
        my_name = my_name_list[len(my_name_list) - 1]
        my_name_format = my_name.split('.')[1]

        newName = str(folderPath) + str('/') + str(name) + str('.') + my_name_format
        os.rename(str(folderPath) + str('/') + str(my_name), newName)
        data[Const.PATH_TO_FILE] = newName

        self.mySheet.UpdateWorkSheetTest(data=data,courseId=data[Const.COURSE_ID],nameOfWork=data[Const.COURSE_ID]+str("TestMidtest"))

    def AddHelpStaff(self,data:{}):
        folderPath = self.mySheet.FindDiractoryPathHelpStaff(data[Const.COURSE_ID], data)
        copy2(data[Const.PATH_TO_FILE], folderPath)

        name = " {} Updated to Semster {} Year {}  Autor by {}".format(
            data[Const.WHAT_THIS_FILE],data[Const.SEMSTER],data[Const.YEAR],
            data[Const.REMARKS])

        my_name_list = data[Const.PATH_TO_FILE].split("\\")
        my_name = my_name_list[len(my_name_list) - 1]
        my_name_format = my_name.split('.')[1]

        newName = str(folderPath) + str('/') + str(name) + str('.') + my_name_format
        os.rename(str(folderPath) + str('/') + str(my_name), newName)
        data[Const.PATH_TO_FILE] = newName

        self.mySheet.UpdateWorkSheetHelpStaff(data=data, courseId=data[Const.COURSE_ID],
                                          nameOfWork=data[Const.COURSE_ID] + str("HelpStaff"))

    def AddHwSol(self,data:{}):
        folderPath = self.mySheet.FindDiractoryPathHW(data[Const.COURSE_ID], data)


        # self.mySheet.UpdateWorkSheet(data[Const.COURSE_ID],data)
        # folderPath='C:/Users/Orr/Desktop/try'
        copy2(data[Const.PATH_TO_FILE], folderPath)

        name = "{} Number {} Part {} Semster {} Year {} LectureName {} Autor by {}".format(
                               data[Const.SOL_OR_HW], data[Const.NUM_OF_HW],
                                data[Const.PART_OF_HW], data[Const.SEMSTER]
                                , data[Const.YEAR], data[Const.LECTURE_NAME], data[Const.REMARKS])

        #sliptingg
        my_name_list=data[Const.PATH_TO_FILE].split("\\")
        my_name=my_name_list[len(my_name_list)-1]
        my_name_format=my_name.split('.')[1]

        newName=str(folderPath)+str('/')+str(name)+str('.')+my_name_format
        os.rename(str(folderPath)+str('/')+str(my_name),newName)
        data[Const.PATH_TO_FILE]=newName

        self.mySheet.UpdateWorkSheetHWSOL(data=data,courseId=data[Const.COURSE_ID],nameOfWork=data[Const.COURSE_ID]+str("HWSolution"))

      #  self.mySheet.FindRootDiractoryID(data["CourseID"])

        pass


    def AddCourse(self,courseName:str,courseId:int):
        self.course_name=courseName
        self.course_id=courseId
        print(os.path.exists(self.working_dirctory))

        #create course deractory
        self.root_diractory=self.working_dirctory+"/{}Id{}".format(self.course_name,str(self.course_id))
        self.mySheet.InsertCourse(self.root_diractory, courseId, courseName)
        os.mkdir(self.root_diractory)

        # create subdiractoryes

        # HELP STAFF
        self.helf_staff_diractory =  self.root_diractory+"/{}".format("HelpStaff")
        os.mkdir(self.helf_staff_diractory)

        # hWSOLUTION

        self.hw_sol_diractory = self.root_diractory + "/{}".format("HwSolution")

        os.mkdir(self.hw_sol_diractory)

        # Lecture

        self.lecture_diractory = self.root_diractory + "/{}".format("Lecture")

        os.mkdir(self.lecture_diractory)

        #Tutorial

        self.tutorial_diractory = self.root_diractory + "/{}".format("Toturial")

        os.mkdir(self.tutorial_diractory)


        # test driactory
        self.test_diractory =  self.root_diractory + "/{}".format("Test")

        os.mkdir(self.test_diractory)

        # test without

        self.test_without_diractory = self.test_diractory + "/{}".format("TestWithOutSol")

        os.mkdir(self.test_without_diractory)

        # test with

        self.test_with_diractory = self.test_diractory + "/{}".format("TestWithSol")

        os.mkdir(self.test_with_diractory)



        #MidTest diractory

        self.mid_test_diractory = self.root_diractory + "/{}".format("MidTest")

        os.mkdir(self.mid_test_diractory)


        self.mid_test_with =  self.mid_test_diractory + "/{}".format("MidTestWithSol")

        os.mkdir(self.mid_test_with)

        self.mid_test_with_out = self.mid_test_diractory+"/{}".format("MidTestWithOutSol")

        os.mkdir(self.mid_test_with_out)






        list_folder_name=[
            'RootFolder',
            'HelpStaffFolder',
            'HWSolutionFolder',
        'LectureFolder',
        'ToturialFolder',
        'TestRootFolder',
        'TestWithOutSolFolder',
        'TestWithSolFolder',
        'MidTestFolder',
        'MidTestWithFolder',
        'MidTestWithOutFolder'
            ]

        list_folder_path=[
            self.root_diractory,
            self.helf_staff_diractory,
            self.hw_sol_diractory,
            self.lecture_diractory,
            self.tutorial_diractory,
            self.test_diractory,
            self.test_without_diractory,
            self.test_with_diractory,
            self.mid_test_diractory,
            self.mid_test_with,
            self.mid_test_with_out

        ]



        course_folder_pd = pd.DataFrame({'FolderName':list_folder_name,
                                         'FolderPath':list_folder_path
                                         })





        #create folder web
        self.mySheet.addNewWorkSheet(course_folder_pd,str(courseId)+'folderId')

        df_list = []



        data_lecture_totorial = {Const.FOLDERID: [],
                                 Const.FILEID: [],
                                 Const.YEAR: [],
                                 Const.SEMSTER: [],
                                 Const.LECTURE_NAME: [],
                                 Const.LECTURE_OR_TOTURIAL: [],
                                 Const.NUMOFLECTURE: [],
                                 Const.PARTOFLECTURE: [],
                                 Const.PATH_TO_FILE_DRIVE: [],
                                 Const.REMARKS: []
                                 }
        df_lecture_totorial = pd.DataFrame(data_lecture_totorial)

        df_list.append((df_lecture_totorial, Const.Toturial_LECTURE_WORK))

        data_HW = {Const.FOLDERID: [],
                   Const.FILEID: [],
                   Const.YEAR: [],
                   Const.SEMSTER: [],
                   Const.LECTURE_NAME: [],
                   Const.SOL_OR_HW: [],
                   Const.NUM_OF_HW: [],
                   Const.PART_OF_HW: [],
                   Const.PATH_TO_FILE_DRIVE: [],
                   Const.REMARKS: []
                   }

        df_HW = pd.DataFrame(data_HW)

        df_list.append((df_HW, Const.HWFOLDER))
        # add חומרי עזר נוספ;

        data_help_staff = {Const.FOLDERID: [],
                           Const.FILEID: [],
                           Const.YEAR: [],
                           Const.SEMSTER: [],
                           Const.WORKSPACE: [],
                           Const.REMARKS: []
                           }

        df_help_staf = pd.DataFrame(data_help_staff)
        df_list.append((df_help_staf, Const.HELPSTAFF))

        data_exam_midexam = {Const.FOLDERID: [],
                                 Const.FILEID: [],
                                 Const.YEAR: [],
                                 Const.SEMSTER: [],
                                 Const.LECTURE_NAME: [],
                                 Const.MOED:[],
                                 Const.TEST_OR_MIDTEST: [],
                                 Const.QUESTION_OR_SOL:[],
                                 Const.PATH_TO_FILE_DRIVE: [],
                                 Const.REMARKS: []
                                 }
        df_exam_midexam = pd.DataFrame(data_exam_midexam)

        df_list.append((df_exam_midexam, Const.TEST_MIDTEST_WORK))



        create = lambda df, name: self.mySheet.addNewWorkSheet(data=df, name_of_work=str(courseId) + str(name))
        [create(item[0],item[1]) for item in df_list]




        try:
            os.mkdir(self.working_dirctory)
        except OSError:
            print("Creation of the directory %s failed" % courseName+str(courseId))
        else:
            print("Successfully created the directory %s " % courseName+str(courseId))






if __name__ == '__main__':
#    os.mkdir('E:/onedriveNew/OneDrive - Technion/CsDriveNew/hello')
   # os.makedirs("exampledirectory")

   # print("Directory '%s' created" % directory)
    one=OneDriveApi("E:/onedriveNew/OneDrive - Technion/CsDriveNew")

    data={
    Const.WHAT_THIS_FILE:"HW",
    Const.NUMOFLECTURE:"1",
    Const.PARTOFLECTURE:"1",
    Const.SEMSTER:"4",
    Const.YEAR:"2012",
    Const.LECTURE_NAME:"Aviv",
    Const.HWSOLUTION:"HW",
    Const.NUM_OF_HW:"1",
    Const.PART_OF_HW:"1",
    Const.SOL_OR_HW:"HW",
    Const.REMARKS:"Orr",
    Const.COURSE_ID:"125345",
    Const.PATH_TO_FILE:"NULL",
    Const.FILEID:"1234"

    }
    #one.AddCourse("infi121",125345)
    name = "{} Number {} Part {} Semster {} Year {} LectureName {} Autor by {}".format(
        data[Const.WHAT_THIS_FILE], data[Const.NUMOFLECTURE],
        data[Const.PARTOFLECTURE], data[Const.SEMSTER]
            , data[Const.YEAR], data[Const.LECTURE_NAME], data[Const.REMARKS])

    one.mySheet.UpdateWorkSheet(data=data, courseId=data[Const.COURSE_ID],
                             nameOfWork=data[Const.COURSE_ID] + str("HWSolution"))









